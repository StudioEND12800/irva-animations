import os
import json
import io
import base64
import tempfile
from datetime import datetime
from flask import Blueprint, send_file, current_app, abort, request, redirect, url_for, session
from models import db, CompteRendu, Photo

export_bp = Blueprint('export', __name__)


def generate_pdf(cr):
    """Génère le PDF du compte-rendu et retourne le chemin du fichier."""
    temp_paths = []
    try:
        from xhtml2pdf import pisa
        from flask import render_template

        pdf_assets = _prepare_pdf_assets(cr, temp_paths)
        html = render_template(
            'dashboard/pdf_cr.html',
            cr=cr,
            data=_deserialize(cr),
            pdf_assets=pdf_assets,
        )
        filename = f'CR_{cr.id}_{_slug(cr.nom_magasin or "inconnu")}.pdf'
        fpath = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)

        with open(fpath, 'wb') as f:
            status = pisa.CreatePDF(html, dest=f, encoding='utf-8')
        if status.err:
            current_app.logger.error('PDF generation returned %s error(s) for CR #%s', status.err, cr.id)
            if os.path.exists(fpath):
                os.remove(fpath)
            return None
        return filename
    except Exception as e:
        current_app.logger.exception('PDF generation error for CR #%s: %s', cr.id, e)
        return None
    finally:
        for path in temp_paths:
            try:
                if path and os.path.exists(path):
                    os.remove(path)
            except OSError:
                current_app.logger.warning('Temporary PDF asset could not be removed: %s', path)


@export_bp.route('/pdf/<int:cr_id>')
def pdf(cr_id):
    if not session.get('admin'):
        return redirect(url_for('dashboard.login', next=request.path))

    cr = CompteRendu.query.get_or_404(cr_id)

    existing_fpath = None
    if cr.pdf_path:
        candidate = os.path.join(current_app.config['UPLOAD_FOLDER'], cr.pdf_path)
        if os.path.exists(candidate):
            existing_fpath = candidate

    # Régénérer à chaque téléchargement admin pour refléter l'état courant du CR.
    pdf_name = generate_pdf(cr)
    if pdf_name:
        if cr.pdf_path != pdf_name:
            cr.pdf_path = pdf_name
            db.session.commit()
        fpath = os.path.join(current_app.config['UPLOAD_FOLDER'], pdf_name)
        return send_file(fpath, as_attachment=True,
                         download_name=f'CR_{cr.nom_magasin}.pdf',
                         mimetype='application/pdf')
    if existing_fpath:
        return send_file(existing_fpath, as_attachment=True,
                         download_name=f'CR_{cr.nom_magasin}.pdf',
                         mimetype='application/pdf')
    abort(500)


@export_bp.route('/excel')
def excel():
    if not session.get('admin'):
        return redirect(url_for('dashboard.login', next=request.full_path.rstrip('?')))

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    filiere = request.args.get('filiere', '')
    annee = request.args.get('annee', '')

    q = CompteRendu.query.filter_by(statut='soumis')
    if filiere:
        q = q.filter_by(filiere=filiere)
    if annee:
        from sqlalchemy import extract
        q = q.filter(extract('year', CompteRendu.date_premier_jour) == int(annee))
    crs = q.order_by(CompteRendu.date_premier_jour).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Décharge données'

    headers = [
        "Date envoi", "Date 1er jour animation", "Nom / Prénom", "N° Cheptel",
        "Animation", "Co-éleveur·se", "Email", "Filière",
        "Enseigne", "Magasin", "CP", "Commune", "Région",
        "Chef boucher", "Ancienneté chef boucher",
        "Rayons présents", "Barquettes LS", "Barquettes sur place",
        "Visibilité LS", "Linéaire LS (m)", "Qualité découpe LS",
        "Outils com LS", "Autre veau LS", "Marque autre veau LS",
        "Visibilité Trad", "Linéaire Trad (m)", "Qualité découpe Trad",
        "Outils com Trad", "Autre veau Trad",
        "Morceaux présents",
        "VAS Escalope €/kg", "VAS Sauté €/kg", "VAS Rôti €/kg",
        "VAS Tendron €/kg", "VAS Jarret €/kg", "VAS Haché €/kg",
        "Autre Escalope €/kg", "Autre Sauté €/kg", "Autre Rôti €/kg",
        "Autre Tendron €/kg", "Autre Jarret €/kg", "Autre Haché €/kg",
        "Prix moy VAS", "Prix moy Autre",
        "Date dernier jour", "Emplacement", "Fréquentation",
        "Approvisionnement", "Ruptures",
        "Outils animation", "Mise en avant magasin",
        "Ventes supplémentaires", "Incident",
        "Tranche âge clients", "Attitude clients",
        "Clients connaissaient VAS", "Type questions",
        "Ressenti mise en place", "Ressenti accroche", "Ressenti argumentaire",
        "Kit IRVA", "Kit Interbev",
        "Échanges chef boucher", "Incident majeur",
        "Statut",
    ]

    # En-tête
    header_fill = PatternFill(start_color='4F7942', end_color='4F7942', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        ws.column_dimensions[cell.column_letter].width = max(12, min(len(h), 25))

    ws.row_dimensions[1].height = 40

    def jl(s):
        if not s:
            return ''
        try:
            return ', '.join(json.loads(s))
        except Exception:
            return str(s)

    for row_num, cr in enumerate(crs, 2):
        row = [
            cr.submitted_at.strftime('%d/%m/%Y %H:%M') if cr.submitted_at else '',
            cr.date_premier_jour.strftime('%d/%m/%Y') if cr.date_premier_jour else '',
            cr.nom_prenom or '',
            cr.num_cheptel or '',
            cr.animation_solo or '',
            cr.nom_coeleveuse or '',
            cr.email or '',
            cr.filiere or '',
            cr.enseigne or '',
            cr.nom_magasin or '',
            cr.code_postal or '',
            cr.commune or '',
            cr.region or '',
            cr.nom_chef_boucher or '',
            cr.anciennete_chef_boucher or '',
            jl(cr.rayons_presents),
            jl(cr.ls_barquettes),
            cr.ls_barquettes_sur_place or '',
            cr.ls_visibilite or '',
            cr.ls_lineaire or '',
            cr.ls_qualite_decoupe or '',
            jl(cr.ls_outils_com),
            cr.ls_autre_veau or '',
            cr.ls_autre_veau_marque or '',
            cr.trad_visibilite or '',
            cr.trad_lineaire or '',
            cr.trad_qualite_decoupe or '',
            jl(cr.trad_outils_com),
            cr.trad_autre_veau or '',
            jl(cr.morceaux_presents),
            cr.prix_vas_escalope or '',
            cr.prix_vas_saute or '',
            cr.prix_vas_roti or '',
            cr.prix_vas_tendron or '',
            cr.prix_vas_jarret or '',
            cr.prix_vas_hache or '',
            cr.prix_autre_escalope or '',
            cr.prix_autre_saute or '',
            cr.prix_autre_roti or '',
            cr.prix_autre_tendron or '',
            cr.prix_autre_jarret or '',
            cr.prix_autre_hache or '',
            cr.prix_moyen_vas or '',
            cr.prix_moyen_autre or '',
            cr.date_dernier_jour.strftime('%d/%m/%Y') if cr.date_dernier_jour else '',
            jl(cr.emplacement_animation),
            cr.frequentation or '',
            cr.approvisionnement or '',
            cr.ruptures or '',
            jl(cr.outils_animation),
            jl(cr.mise_en_avant),
            cr.ventes_supplementaires or '',
            cr.incident or '',
            jl(cr.tranche_age),
            jl(cr.attitude_clients),
            cr.clients_connaissaient_vas or '',
            jl(cr.type_questions),
            cr.ressenti_mise_en_place or '',
            cr.ressenti_accroche or '',
            cr.ressenti_argumentaire or '',
            cr.kit_irva or '',
            cr.kit_interbev or '',
            jl(cr.echanges_chef_boucher),
            cr.incident_majeur or '',
            cr.statut or '',
        ]
        for col, val in enumerate(row, 1):
            ws.cell(row=row_num, column=col, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    now = datetime.now().strftime('%Y%m%d')
    fname = f'CR_Animations_IRVA_{now}.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def _deserialize(cr):
    def jl(s):
        if not s:
            return []
        try:
            return json.loads(s)
        except Exception:
            return [s] if s else []
    return {k: jl(getattr(cr, k)) for k in [
        'rayons_presents', 'ls_barquettes', 'ls_outils_com', 'trad_outils_com',
        'morceaux_presents', 'emplacement_animation', 'outils_animation',
        'mise_en_avant', 'tranche_age', 'attitude_clients', 'type_questions',
        'echanges_chef_boucher', 'type_incident',
    ]}


def _slug(s):
    import re
    return re.sub(r'[^a-z0-9]+', '_', s.lower())[:40]


def _prepare_pdf_assets(cr, temp_paths):
    assets = {
        'signature_eleveur_uri': None,
        'signature_boucher_uri': None,
    }

    if cr.signature_eleveur_data:
        temp_file = _write_data_url_image(cr.signature_eleveur_data, f'sig_eleveur_{cr.id}_')
        if temp_file:
            temp_paths.append(temp_file)
            assets['signature_eleveur_uri'] = os.path.abspath(temp_file)

    if cr.signature_boucher_path:
        signature_path = os.path.join(current_app.config['UPLOAD_FOLDER'], cr.signature_boucher_path)
        if os.path.exists(signature_path):
            assets['signature_boucher_uri'] = os.path.abspath(signature_path)

    return assets


def _write_data_url_image(data_url, prefix):
    if not data_url or ',' not in data_url:
        return None

    header, encoded = data_url.split(',', 1)
    extension = _image_extension_from_header(header)

    try:
        binary = base64.b64decode(encoded, validate=False)
    except Exception:
        current_app.logger.warning('Invalid signature data URL for PDF generation')
        return None

    fd, temp_path = tempfile.mkstemp(prefix=prefix, suffix=extension)
    try:
        with os.fdopen(fd, 'wb') as f:
            f.write(binary)
    except Exception:
        try:
            os.close(fd)
        except OSError:
            pass
        if os.path.exists(temp_path):
            os.remove(temp_path)
        raise
    return temp_path


def _image_extension_from_header(header):
    header = header.lower()
    if 'image/jpeg' in header or 'image/jpg' in header:
        return '.jpg'
    if 'image/gif' in header:
        return '.gif'
    return '.png'
