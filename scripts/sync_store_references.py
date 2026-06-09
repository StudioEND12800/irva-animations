#!/usr/bin/env python3

from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from app.store_reference import sync_store_reference
from app.utils import compact_spaces, infer_department_code, infer_region
from models import CompteRendu, db
from wsgi import app


def _clean_cell(value) -> str:
    return compact_spaces(value)


def import_workbook(workbook_path: Path, sheet_name: str) -> dict[str, int]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise SystemExit(f"Feuille introuvable : {sheet_name}")

    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(min_row=2, values_only=True)
    stats = {'created': 0, 'updated': 0, 'rows': 0}

    for row in rows:
        canonical_name = _clean_cell(row[1] if len(row) > 1 else '')
        if not canonical_name:
            continue
        aliases = [_clean_cell(cell) for cell in row[2:] if _clean_cell(cell)]
        _reference, created = sync_store_reference(
            nom_magasin=canonical_name,
            aliases=aliases,
        )
        stats['created' if created else 'updated'] += 1
        stats['rows'] += 1

    workbook.close()
    return stats


def merge_existing_forms() -> dict[str, int]:
    stats = {'created': 0, 'updated': 0, 'rows': 0}
    records = CompteRendu.query.filter(CompteRendu.nom_magasin.isnot(None)).all()
    for cr in records:
        nom_magasin = compact_spaces(cr.nom_magasin)
        if not nom_magasin:
            continue
        code_postal = compact_spaces(cr.code_postal)
        code_departement = compact_spaces(cr.code_departement) or infer_department_code(code_postal)
        region = compact_spaces(cr.region) or infer_region(code_postal, code_departement)
        _reference, created = sync_store_reference(
            enseigne=compact_spaces(cr.enseigne),
            nom_magasin=nom_magasin,
            code_postal=code_postal,
            commune=compact_spaces(cr.commune),
            code_departement=code_departement,
            region=region,
            aliases=[nom_magasin],
        )
        stats['created' if created else 'updated'] += 1
        stats['rows'] += 1
    return stats


def main():
    parser = argparse.ArgumentParser(description="Synchronise le référentiel magasins IRVA.")
    parser.add_argument('--workbook', type=Path, help="Chemin du fichier Excel source.")
    parser.add_argument('--sheet', default='Liste nom mag', help="Nom de la feuille à importer.")
    parser.add_argument('--merge-forms', action='store_true', help="Enrichit aussi le référentiel avec les formulaires déjà reçus.")
    parser.add_argument('--dry-run', action='store_true', help="Ne valide aucune écriture en base.")
    args = parser.parse_args()

    if not args.workbook and not args.merge_forms:
        raise SystemExit("Indiquez --workbook et/ou --merge-forms.")

    with app.app_context():
        workbook_stats = {'created': 0, 'updated': 0, 'rows': 0}
        form_stats = {'created': 0, 'updated': 0, 'rows': 0}

        if args.workbook:
            workbook_stats = import_workbook(args.workbook, args.sheet)

        if args.merge_forms:
            form_stats = merge_existing_forms()

        if args.dry_run:
            db.session.rollback()
        else:
            db.session.commit()

    print(
        f"workbook_rows={workbook_stats['rows']} workbook_created={workbook_stats['created']} workbook_updated={workbook_stats['updated']}"
    )
    print(
        f"form_rows={form_stats['rows']} form_created={form_stats['created']} form_updated={form_stats['updated']}"
    )
    print(f"dry_run={'yes' if args.dry_run else 'no'}")


if __name__ == '__main__':
    main()
